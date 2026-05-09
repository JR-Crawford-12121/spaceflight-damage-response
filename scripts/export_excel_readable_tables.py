#!/usr/bin/env python3
"""
Export main pipeline CSV tables to a formatted Excel workbook (readability only).

Does not modify analysis outputs — reads CSVs and writes a separate .xlsx file.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
TABLES = ROOT / "outputs" / "tables"
OUTPUT_XLSX = TABLES / "readable_results_workbook.xlsx"

CSV_ORDER: list[str] = [
    "main_candidate_signals.csv",
    "pathway_scores_evidence_filtered.csv",
    "top_genes_from_pathway_level_signals.csv",
    "top_damage_response_genes.csv",
]

OPTIONAL_CSV = "attention_items.csv"

LONG_TEXT_COLUMNS = frozenset(
    {
        "reason_to_notice",
        "caution",
        "interpretation_label",
        "top_contributing_genes",
    }
)

MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 45
LONG_TEXT_MIN_WIDTH = 14


def _unique_sheet_name(base: str, used: set[str]) -> str:
    """Excel sheet names are max 31 characters and must be unique."""
    s = base[:31]
    if s not in used:
        used.add(s)
        return s
    n = 2
    while True:
        suffix = f"_{n}"
        trimmed = base[: 31 - len(suffix)] + suffix
        if trimmed not in used:
            used.add(trimmed)
            return trimmed[:31]
        n += 1


def _column_width(header: str | None, column_values: list, column_key: str) -> float:
    lengths: list[int] = [len(str(header)) if header is not None else 0]
    for v in column_values:
        if v is None:
            continue
        try:
            if pd.isna(v):
                continue
        except TypeError:
            pass
        lengths.append(len(str(v)))
    longest = max(lengths) if lengths else MIN_COL_WIDTH
    padded = min(longest + 2, MAX_COL_WIDTH)
    floor = LONG_TEXT_MIN_WIDTH if column_key in LONG_TEXT_COLUMNS else MIN_COL_WIDTH
    return float(min(MAX_COL_WIDTH, max(floor, padded)))


def _apply_formatting(path: Path) -> None:
    wb = load_workbook(path)
    top_align_wrap = Alignment(wrap_text=True, vertical="top")
    bold = Font(bold=True)

    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column
        if max_col < 1:
            continue

        last_letter = get_column_letter(max_col)
        ws.freeze_panes = "A2"
        if max_row >= 1:
            ws.auto_filter.ref = f"A1:{last_letter}{max_row}"

        for col_idx in range(1, max_col + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            col_key = str(header_cell.value) if header_cell.value is not None else ""
            header_cell.font = bold
            header_cell.alignment = top_align_wrap

            col_values = [
                ws.cell(row=r, column=col_idx).value for r in range(2, max_row + 1)
            ]
            width = _column_width(header_cell.value, col_values, col_key)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

            for r in range(2, max_row + 1):
                c = ws.cell(row=r, column=col_idx)
                c.alignment = top_align_wrap

    wb.save(path)


def main() -> int:
    TABLES.mkdir(parents=True, exist_ok=True)

    inputs: list[tuple[Path, str]] = []
    used_names: set[str] = set()

    for name in CSV_ORDER:
        p = TABLES / name
        if not p.is_file():
            print(f"Warning: missing {p.relative_to(ROOT)} — skipped.", file=sys.stderr)
            continue
        sn = _unique_sheet_name(Path(name).stem, used_names)
        inputs.append((p, sn))

    opt = TABLES / OPTIONAL_CSV
    if opt.is_file():
        sn = _unique_sheet_name(opt.stem, used_names)
        inputs.append((opt, sn))

    if not inputs:
        print("No input CSVs found; nothing to write.", file=sys.stderr)
        return 1

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        for csv_path, sheet in inputs:
            df = pd.read_csv(csv_path)
            df.to_excel(writer, sheet_name=sheet, index=False)

    _apply_formatting(OUTPUT_XLSX)
    print("Wrote outputs/tables/readable_results_workbook.xlsx")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
