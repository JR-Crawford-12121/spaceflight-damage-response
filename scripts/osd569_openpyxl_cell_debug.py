"""
Low-level openpyxl dump for GLDS-561 / I4-LP2 (merged cells, raw strings).

Writes: outputs/logs/osd569_openpyxl_cell_debug.txt
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

import pandas as pd

from load_data import OSD569_FILENAME, locate_file  # noqa: E402
from osd569_parse import find_two_row_header_indices  # noqa: E402


def main() -> int:
    out_path = ROOT / "outputs" / "logs" / "osd569_openpyxl_cell_debug.txt"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []

    xlsx = locate_file(OSD569_FILENAME, ROOT)
    if xlsx is None:
        msg = f"File not found: {OSD569_FILENAME}"
        out_path.write_text(msg, encoding="utf-8")
        print(msg)
        return 1

    sheet_name = "I4-LP2"
    preview = pd.read_excel(xlsx, sheet_name=sheet_name, header=None, nrows=45, engine="openpyxl")
    h0 = find_two_row_header_indices(preview)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        lines.append(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames!r}")
        out_path.write_text("\n".join(lines), encoding="utf-8")
        return 1

    ws = wb[sheet_name]
    lines.append(f"workbook: {xlsx}")
    lines.append(f"sheet: {sheet_name}")
    lines.append(f"pandas-detected first header row index (0-based in preview): {h0}")
    lines.append(f"(Excel row number for that row ≈ {h0 + 1} if sheet starts at row 1)")
    lines.append("")

    lines.append("=== Rows 1–20 (Excel 1-based), columns A–N ===")
    for r in range(1, 21):
        row_parts = []
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            letter = get_column_letter(c)
            v = cell.value
            row_parts.append(f"{letter}{r}={v!r}")
        lines.append(" | ".join(row_parts))

    lines.append("")
    lines.append("=== Header region rows 8–14, columns A–M (inspect merged-header overlap) ===")
    for r in range(8, 15):
        row_parts = []
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            letter = get_column_letter(c)
            row_parts.append(f"{letter}{r}={cell.value!r}")
        lines.append(" | ".join(row_parts))

    lines.append("")
    lines.append(
        "=== Probe columns often labeled DESeq2 / pipeline near detected header ==="
    )
    header_row = (h0 + 1) if h0 is not None else 11
    target_tokens = ("log2", "DESeq2", "pipeline", "p-value", "adjusted")
    col_hits: list[tuple[int, str]] = []
    for c in range(1, min(ws.max_column + 1, 40)):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        sv = str(v).lower()
        if any(t.lower() in sv for t in target_tokens):
            col_hits.append((c, str(v)))

    lines.append(f"cells on Excel row {header_row} matching tokens (first 40 cols scanned):")
    for c, txt in col_hits[:25]:
        letter = get_column_letter(c)
        lines.append(f"  {letter}{header_row}: {txt!r}")

    lines.append("")
    lines.append(f"=== First 10 data values under those columns (rows {header_row + 1} .. +10) ===")
    for c, txt in col_hits[:12]:
        letter = get_column_letter(c)
        vals = []
        for rr in range(header_row + 1, header_row + 11):
            vals.append(ws.cell(row=rr, column=c).value)
        lines.append(f"  {letter} ({txt[:40]!r}): {vals}")

    wb.close()
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote: outputs/logs/osd569_openpyxl_cell_debug.txt")
    print(f"Full path: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
